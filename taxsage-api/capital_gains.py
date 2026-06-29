import io
import json
import base64

import pikepdf
import openpyxl
import anthropic

STCG_RATE = 0.20
LTCG_RATE = 0.125
LTCG_EXEMPTION = 125000

# --- Claude prompts ---

KUVERA_SYSTEM = "You are a financial document parser. Extract equity capital gains totals from a Kuvera Capital Gains Statement."
# Kuvera's combined "Total" includes equity + debt — must use "Equity Sub Total" to avoid debt mixing
KUVERA_PROMPT = """From this Kuvera Capital Gains Statement PDF, extract TWO separate totals from the SUMMARY section only (not per-fund rows).

1. EQUITY gains: use the "Equity Sub Total" line — NOT the combined "Total".
2. DEBT gains: use the "Debt Sub Total" line — this is gains from debt mutual funds.

Return ONLY this JSON, no other text:
{"stcg_equity": <integer rupees>, "ltcg_equity": <integer rupees>, "stcg_debt": <integer rupees>, "ltcg_debt": <integer rupees>}
All values 0 if not present."""

CAMS_SYSTEM = "You are a financial document parser. Extract equity capital gains totals from a CAMS Capital Gain/Loss Statement."
CAMS_PROMPT = """From the "Capital Gain/Loss \u2013 Overall Summary (Equity)" section ONLY.
Extract Short Term Capital Gain total and Long Term Capital Gain WITHOUT indexation total.
Ignore Non Equity section entirely.
Return ONLY this JSON, no other text:
{"stcg_total": <integer rupees, 0 if none>, "ltcg_total": <integer rupees from without-indexation column, 0 if none>}"""

GENERIC_SYSTEM = "You are a financial document parser. Extract equity capital gains from a Kuvera or CAMS capital gains statement."
GENERIC_PROMPT = """Extract equity capital gains from the summary section only.
If CAMS: use "Capital Gain/Loss \u2013 Overall Summary (Equity)", use without-indexation column for LTCG. Ignore Non Equity.
If Kuvera: use top summary totals, not per-fund rows.
Return ONLY this JSON, no other text:
{"stcg_total": <integer rupees, 0 if none>, "ltcg_total": <integer rupees, 0 if none>}"""


def _decrypt_pdf(pdf_bytes: bytes, password: str) -> bytes:
    """Open PDF — if encrypted, decrypt with password. Raises ValueError('wrong_password') on failure."""
    try:
        with pikepdf.open(io.BytesIO(pdf_bytes)) as pdf:
            buf = io.BytesIO()
            pdf.save(buf)
            return buf.getvalue()
    except pikepdf.PasswordError:
        pass
    try:
        with pikepdf.open(io.BytesIO(pdf_bytes), password=password) as pdf:
            buf = io.BytesIO()
            pdf.save(buf)
            return buf.getvalue()
    except pikepdf.PasswordError:
        raise ValueError("wrong_password")


def _call_claude(pdf_bytes: bytes, system: str, prompt: str, expected_keys=None) -> dict:
    """Send PDF to Claude, extract JSON. Retries once on bad shape — fails loudly rather than returning zeros."""
    if expected_keys is None:
        expected_keys = ["stcg_total", "ltcg_total"]
    client = anthropic.Anthropic()
    b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

    def _attempt():
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=512,
            system=system,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "document",
                     "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        raw = resp.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()
        data = json.loads(raw)
        if not all(k in data for k in expected_keys):
            raise ValueError("unexpected_shape")
        return data

    try:
        data = _attempt()
    except (json.JSONDecodeError, ValueError, anthropic.APIError):
        try:
            data = _attempt()
        except (json.JSONDecodeError, ValueError, anthropic.APIError) as e:
            raise ValueError(f"claude_parse_failed: {e}")

    return {k: int(data[k] or 0) for k in expected_keys}


def _parse_zerodha(xlsx_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    stcg, ltcg = 0.0, 0.0

    if "Equity and Non Equity" in wb.sheetnames:
        for row in wb["Equity and Non Equity"].iter_rows():
            if len(row) < 3 or row[1].value is None:
                continue
            label = str(row[1].value).strip()
            val = row[2].value
            if val is None:
                continue
            if label == "Short Term profit":
                stcg += float(val)
            elif label == "Long Term profit":
                ltcg += float(val)

    if "Mutual Funds" in wb.sheetnames:
        for row in wb["Mutual Funds"].iter_rows():
            if len(row) < 3 or row[1].value is None:
                continue
            label = str(row[1].value).strip()
            val = row[2].value
            if val is None:
                continue
            if label == "Short Term profit Equity":
                stcg += float(val)
            elif label == "Long Term profit Equity":
                ltcg += float(val)

    return {"stcg": int(stcg), "ltcg": int(ltcg)}


def _parse_pdf(pdf_bytes: bytes, password: str, filename: str) -> dict:
    decrypted = _decrypt_pdf(pdf_bytes, password)
    fname = filename.lower()
    if "kuvera" in fname:
        result = _call_claude(decrypted, KUVERA_SYSTEM, KUVERA_PROMPT,
                              expected_keys=["stcg_equity", "ltcg_equity", "stcg_debt", "ltcg_debt"])
        return {"stcg": result["stcg_equity"], "ltcg": result["ltcg_equity"],
                "debt_stcg": result["stcg_debt"], "debt_ltcg": result["ltcg_debt"]}
    elif "cams" in fname:
        result = _call_claude(decrypted, CAMS_SYSTEM, CAMS_PROMPT)
    else:
        result = _call_claude(decrypted, GENERIC_SYSTEM, GENERIC_PROMPT)
    return {"stcg": result["stcg_total"], "ltcg": result["ltcg_total"], "debt_stcg": 0, "debt_ltcg": 0}


def process(files: list, password: str) -> dict:
    """files: list of (filename, bytes). Returns combined capital gains summary."""
    net_stcg, net_ltcg, net_debt_ltcg = 0, 0, 0
    sources_parsed = []

    for filename, file_bytes in files:
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext == "xlsx":
            result = _parse_zerodha(file_bytes)
            source_type = "zerodha_xlsx"
        elif ext == "pdf":
            result = _parse_pdf(file_bytes, password, filename)
            fname = filename.lower()
            source_type = "kuvera_pdf" if "kuvera" in fname else "cams_pdf" if "cams" in fname else "pdf"
        else:
            raise ValueError(f"unsupported_file:{filename}")

        net_stcg += result["stcg"]
        net_ltcg += result["ltcg"]
        net_debt_ltcg += result.get("debt_ltcg", 0)
        sources_parsed.append({"filename": filename, "type": source_type,
                                "stcg": result["stcg"], "ltcg": result["ltcg"],
                                "debt_ltcg": result.get("debt_ltcg", 0)})

    # Deduplication: if both Kuvera and CAMS PDFs present, CAMS equity is a subset of Kuvera
    kuvera_entry = next((s for s in sources_parsed if s["type"] == "kuvera_pdf"), None)
    cams_entry = next((s for s in sources_parsed if s["type"] == "cams_pdf"), None)
    dedup_applied = False
    dedup_amount = 0
    if kuvera_entry and cams_entry:
        dedup_amount = cams_entry["ltcg"]
        net_ltcg -= dedup_amount
        net_stcg -= cams_entry["stcg"]
        dedup_applied = True
        cams_entry["note"] = "Deduplicated — transactions already included in Kuvera report"

    ltcg_taxable = max(0, net_ltcg - LTCG_EXEMPTION)
    # Losses (negative net_stcg) are not taxed — preserve sign in output but apply zero tax
    stcg_tax = int(net_stcg * STCG_RATE) if net_stcg > 0 else 0
    ltcg_tax = int(ltcg_taxable * LTCG_RATE)

    result_dict = {
        "net_stcg": net_stcg,
        "net_ltcg": net_ltcg,
        "ltcg_exemption_used": min(net_ltcg, LTCG_EXEMPTION),
        "ltcg_taxable": ltcg_taxable,
        "stcg_tax": stcg_tax,
        "ltcg_tax": ltcg_tax,
        "total_tax_owed": stcg_tax + ltcg_tax,
        "net_debt_ltcg": net_debt_ltcg,
        "debt_ltcg_note": "Taxed at income slab rate, not 12.5%. Add to Schedule CG B5." if net_debt_ltcg > 0 else None,
        "sources_parsed": sources_parsed,
    }
    if dedup_applied:
        result_dict["dedup_applied"] = True
        result_dict["dedup_amount"] = dedup_amount
    return result_dict
